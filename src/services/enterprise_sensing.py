"""Enterprise evidence lifecycle, file extraction, and strategy eligibility."""

from __future__ import annotations

import csv
import hashlib
import io
from datetime import UTC, datetime

from src.models.enterprise import (
    EnterpriseEvidenceCategory,
    EnterpriseEvidenceItem,
    EnterpriseReviewStatus,
    EnterpriseSensingArtifact,
    EnterpriseSensitivity,
    EnterpriseStatementType,
)
from src.state.project import ProjectState


MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_EXTRACTED_CHARACTERS = 50_000


class EnterpriseSensingError(ValueError):
    pass


def new_enterprise_artifact(project: ProjectState) -> EnterpriseSensingArtifact:
    return EnterpriseSensingArtifact(
        project_id=project.project_id,
        target_company_snapshot=project.target_company,
        strategy_objective_snapshot=project.company_strategy_objective,
    )


def load_redacted_demo_enterprise_pack(
    project: ProjectState,
    artifact: EnterpriseSensingArtifact | None = None,
) -> EnterpriseSensingArtifact:
    """Add fictitious, clearly labelled inputs that demonstrate the extension layer."""

    examples = (
        EnterpriseEvidenceItem(
            title="模拟：核心医院与渠道覆盖基础",
            category=EnterpriseEvidenceCategory.SALES_CHANNEL,
            statement_type=EnterpriseStatementType.OBSERVATION,
            content=(
                "脱敏模拟信息：公司在核心区域拥有成熟的医院客户与经销商触点，但渠道反馈显示，"
                "客户正在从单项试剂采购转向对样本到报告的一体化交付能力进行评价。"
            ),
            source_owner="模拟销售负责人",
            strategic_relevance="用于判断现有渠道资产能否支持目标细分市场进入与解决方案式销售。",
            sensitivity=EnterpriseSensitivity.REDACTED_DEMO,
            input_method="demo_pack",
        ),
        EnterpriseEvidenceItem(
            title="模拟：现有产品与技术能力边界",
            category=EnterpriseEvidenceCategory.RESEARCH_DEVELOPMENT,
            statement_type=EnterpriseStatementType.FACT,
            content=(
                "脱敏模拟信息：公司现有研发与注册资源主要集中在PCR平台；数字PCR和高通量测序"
                "项目尚处早期验证阶段，未来十二个月可并行推进的新增注册项目数量有限。"
            ),
            source_owner="模拟研发负责人",
            strategic_relevance="用于评估进入新技术细分市场的能力差距、资源约束与合作需求。",
            sensitivity=EnterpriseSensitivity.REDACTED_DEMO,
            input_method="demo_pack",
        ),
        EnterpriseEvidenceItem(
            title="模拟：客户未满足需求",
            category=EnterpriseEvidenceCategory.CUSTOMER,
            statement_type=EnterpriseStatementType.OBSERVATION,
            content=(
                "脱敏模拟信息：受访实验室更关注周转时间、自动化、质控与信息系统连接，单纯降低"
                "试剂价格并不足以推动其更换平台；不同等级医院对菜单广度和服务响应要求不同。"
            ),
            source_owner="模拟客户洞察负责人",
            strategic_relevance="用于验证客户需求是否支持从产品销售向工作流解决方案转型。",
            sensitivity=EnterpriseSensitivity.REDACTED_DEMO,
            input_method="demo_pack",
        ),
        EnterpriseEvidenceItem(
            title="模拟：投资与现金流约束",
            category=EnterpriseEvidenceCategory.FINANCE,
            statement_type=EnterpriseStatementType.HYPOTHESIS,
            content=(
                "脱敏模拟假设：管理层要求新业务在分阶段里程碑下投入，若临床验证、注册或渠道"
                "转化未达到预设阈值，应停止继续扩张或改为合作模式。"
            ),
            source_owner="模拟财务负责人",
            strategic_relevance="用于设计行动优先级、资源分配、领先指标和停止条件。",
            sensitivity=EnterpriseSensitivity.REDACTED_DEMO,
            input_method="demo_pack",
        ),
        EnterpriseEvidenceItem(
            title="模拟：管理层战略意图",
            category=EnterpriseEvidenceCategory.STRATEGIC_INTENT,
            statement_type=EnterpriseStatementType.STRATEGIC_INTENT,
            content=(
                "脱敏模拟信息：管理层希望在未来三年形成第二增长曲线，但不接受以大规模低价"
                "换取短期份额，优先考虑可复用现有渠道、可建立临床证据壁垒的机会。"
            ),
            source_owner="模拟管理层",
            strategic_relevance="作为评分中的战略适配基准，并约束Action Plan的选择与排序。",
            sensitivity=EnterpriseSensitivity.REDACTED_DEMO,
            input_method="demo_pack",
        ),
    )
    current = artifact or new_enterprise_artifact(project)
    existing_titles = {item.title for item in current.entries}
    entries = [*current.entries, *(item for item in examples if item.title not in existing_titles)]
    return current.model_copy(
        update={
            "entries": entries,
            "target_company_snapshot": project.target_company,
            "strategy_objective_snapshot": project.company_strategy_objective,
            "human_confirmed": False,
            "confirmed_at": None,
            "updated_at": datetime.now(UTC),
        }
    )


def upsert_enterprise_entry(
    artifact: EnterpriseSensingArtifact | None,
    project: ProjectState,
    item: EnterpriseEvidenceItem,
) -> EnterpriseSensingArtifact:
    current = artifact or new_enterprise_artifact(project)
    return current.model_copy(
        update={
            "entries": [*current.entries, item],
            "target_company_snapshot": project.target_company,
            "strategy_objective_snapshot": project.company_strategy_objective,
            "human_confirmed": False,
            "confirmed_at": None,
            "updated_at": datetime.now(UTC),
        }
    )


def review_enterprise_entry(
    artifact: EnterpriseSensingArtifact,
    enterprise_evidence_id: str,
    status: EnterpriseReviewStatus,
    note: str | None = None,
) -> EnterpriseSensingArtifact:
    if status not in {EnterpriseReviewStatus.ACCEPTED, EnterpriseReviewStatus.REJECTED}:
        raise ValueError("enterprise review can only accept or reject")
    found = False
    entries: list[EnterpriseEvidenceItem] = []
    for item in artifact.entries:
        if item.enterprise_evidence_id == enterprise_evidence_id:
            found = True
            item = item.model_copy(
                update={
                    "review_status": status,
                    "reviewer_note": note.strip() if note and note.strip() else None,
                    "reviewed_at": datetime.now(UTC),
                }
            )
        entries.append(item)
    if not found:
        raise ValueError(f"unknown enterprise evidence: {enterprise_evidence_id}")
    return artifact.model_copy(
        update={
            "entries": entries,
            "human_confirmed": False,
            "confirmed_at": None,
            "updated_at": datetime.now(UTC),
        }
    )


def confirm_enterprise_artifact(
    artifact: EnterpriseSensingArtifact,
    project: ProjectState,
    *,
    consent_to_model_processing: bool,
    public_demo_acknowledged: bool,
) -> EnterpriseSensingArtifact:
    if not artifact.accepted_entries:
        raise EnterpriseSensingError("至少需要一条人工接受的Enterprise Evidence")
    if not consent_to_model_processing:
        raise EnterpriseSensingError("必须明确允许本项目模型处理已接受的企业资料")
    if not public_demo_acknowledged:
        raise EnterpriseSensingError("必须确认公开演示版只使用脱敏或模拟资料")
    if any(
        item.sensitivity != EnterpriseSensitivity.REDACTED_DEMO
        for item in artifact.accepted_entries
    ):
        raise EnterpriseSensingError(
            "公开演示版不能确认未脱敏的内部、机密或受限资料，请改用脱敏或模拟内容"
        )
    return artifact.model_copy(
        update={
            "target_company_snapshot": project.target_company,
            "strategy_objective_snapshot": project.company_strategy_objective,
            "consent_to_model_processing": True,
            "public_demo_acknowledged": True,
            "human_confirmed": True,
            "confirmed_at": datetime.now(UTC),
            "updated_at": datetime.now(UTC),
        }
    )


def company_strategy_gate_reasons(project: ProjectState) -> list[str]:
    reasons: list[str] = []
    if not project.company_strategy_enabled:
        reasons.append("项目未启用企业战略路径")
    if not project.target_company:
        reasons.append("尚未填写目标企业")
    if not project.company_strategy_objective:
        reasons.append("尚未填写企业战略意图或战略目标")
    artifact = project.enterprise_sensing_artifact
    if artifact is None:
        reasons.append("尚未建立Enterprise Sensing资料")
        return reasons
    if artifact.target_company_snapshot != project.target_company:
        reasons.append("Enterprise Sensing中的目标企业快照已过期")
    if artifact.strategy_objective_snapshot != project.company_strategy_objective:
        reasons.append("企业战略目标已变化，需要重新确认Enterprise Sensing")
    if not artifact.accepted_entries:
        reasons.append("尚无人工接受的Enterprise Evidence")
    if not artifact.consent_to_model_processing:
        reasons.append("尚未允许模型在本项目中使用企业资料")
    if not artifact.human_confirmed:
        reasons.append("Enterprise Sensing尚未完成人工确认")
    return reasons


def enterprise_item_from_upload(
    *,
    file_name: str,
    mime_type: str | None,
    data: bytes,
    source_owner: str,
    strategic_relevance: str,
    sensitivity: EnterpriseSensitivity,
) -> EnterpriseEvidenceItem:
    if not data:
        raise EnterpriseSensingError("上传文件为空")
    if len(data) > MAX_UPLOAD_BYTES:
        raise EnterpriseSensingError("单个文件不能超过5MB")
    text = extract_document_text(file_name, mime_type, data)
    if not text.strip():
        raise EnterpriseSensingError("文件未提取到可用文字")
    return EnterpriseEvidenceItem(
        title=file_name,
        category=EnterpriseEvidenceCategory.INTERNAL_DOCUMENT,
        statement_type=EnterpriseStatementType.MIXED_DOCUMENT,
        content=text[:MAX_EXTRACTED_CHARACTERS],
        source_owner=source_owner,
        strategic_relevance=strategic_relevance,
        sensitivity=sensitivity,
        input_method="file",
        file_name=file_name,
        file_sha256=hashlib.sha256(data).hexdigest(),
    )


def extract_document_text(file_name: str, mime_type: str | None, data: bytes) -> str:
    suffix = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""
    if suffix in {"txt", "md", "csv"}:
        decoded = _decode_text(data)
        if suffix == "csv":
            rows = csv.reader(io.StringIO(decoded))
            return "\n".join(" | ".join(cell.strip() for cell in row) for row in rows)
        return decoded
    if suffix == "pdf" or mime_type == "application/pdf":
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(data))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    if suffix == "docx":
        from docx import Document

        document = Document(io.BytesIO(data))
        return "\n".join(paragraph.text for paragraph in document.paragraphs)
    if suffix == "xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"[Sheet: {sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                lines.append(" | ".join("" if value is None else str(value) for value in row))
        return "\n".join(lines)
    raise EnterpriseSensingError("暂不支持该文件格式")


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise EnterpriseSensingError("无法识别文本文件编码")
